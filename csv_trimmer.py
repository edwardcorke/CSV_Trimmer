import csv
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.colors import Color
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
# from datetime import datetime


# Return index in inputted list where inputted value is found (-1 if not found)
def list_index_by_string(list, value):
    for index in range(len(list)):
        if value in list[index]:
            return index
    return -1


# Read headers.config to set what columns to keep
desired_columns = []
headers_config_file = open("headers.config",'r')
splitting_student_name = False
for row in headers_config_file:
    if row.strip() == "Student Name":
        splitting_student_name = True
        desired_columns.append("Firstname")
        desired_columns.append("Lastname")
    else:
        desired_columns.append(row.strip())


# Read in CSV file
csv_data = ""
filename = input("Enter filename: ")
with open(filename) as f:
    csv_data += f.read() + '\n' # add trailing new line character


# Convert into list of strings
csv_data_list = csv_data.split('\n')

# Get headers
headers = csv_data_list[0].encode('latin-1').decode('utf-8')
headers = headers.replace("\"",'').split(',')
headers[0] = headers[0].replace("\ufeff", "")

# Setting filename of file exported
export_filename = "unnamed_export.xlsx"
col_pos = list_index_by_string(headers, "Item Name")
if col_pos != -1:
    export_filename = csv_data_list[1].split(',')[col_pos] + ".xlsx"
    export_filename = export_filename.replace("\"",'')
    export_filename = export_filename.replace("\'",'')

# Split 'Student Name' into 'Firstname' and 'Lastname'
updated_csv = []
# Find the index of 'Student Name' in the headers list
insert_index = list_index_by_string(headers, "Student Name")

# Scan down csv line by line
for index in range(1, len(csv_data_list) - 2):
    line = csv_data_list[index]
    line_list = line.split(',')

    if splitting_student_name:
        #Split 'Student Name' into list separated by space (with max split of 1)
        try:
            firstname, lastname = line_list[insert_index].split(' ', 1)
        except:
            firstname = line_list[insert_index]
            lastname = ""
        # Insert firstname and lastname and remove " to ensure cells are read correctly
        line_list[insert_index] = firstname.replace("\"","")
        line_list.insert(insert_index + 1, lastname.replace("\"",""))
    # Update csv data
    updated_csv.append(line_list)

# Update headers with new columns
if splitting_student_name:
    headers[insert_index] = "Firstname"
    headers.insert(insert_index + 1, "Lastname")
headers[-1] += "\n"

# Convert update csv data into string of strings separated by '\n'
csv_data_new = ""
for item in updated_csv:
    csv_data_new += ','.join(item) + '\n'
# Add headers
csv_data_new = ",".join(headers) + csv_data_new
# Convert data back into list of strings so csv.DictReader can read it
csv_data_new = csv_data_new.split("\n")


# Read data into csv.DictReader
csv_file = csv.DictReader(csv_data_new)

data = []
# Remove any items in desired_columns that aren't in headers
desired_columns_temp = []
# Remove '\n' in last header for correct matching
headers[-1] = headers[-1].replace("\n","")
for column in desired_columns:
    if column in headers:
        desired_columns_temp.append(column)
desired_columns = desired_columns_temp
data.append(desired_columns)
selected_data = []


# Read, select and append data
for row in csv_file:
    for key in desired_columns:
        try:
            selected_data.append(row[key].encode('latin-1').decode('utf-8'))
        except:
            pass
    data.append(selected_data)
    selected_data = []


# Generate Excel Document
# Create new workbook
wb = Workbook()
try:
    wb.save(export_filename)
except:
    print("Cannot write to file- already open")
    exit()

# Load existing workbook
wb = load_workbook(export_filename)
# Activate worksheet to write dataframe
active = wb['Sheet']

# Insert data into each row of active sheet
for row in data:
    active.append(row)


# Styling Excel Sheet
# Style headers
header_bg_colour = PatternFill(start_color='C8C8C8',
                   end_color='C8C8C8',
                   fill_type='solid')

for cell in active[1]:
    cell.font = Font(bold=True)
    cell.fill = header_bg_colour

alternating_row_colour = PatternFill(start_color='F2F2F2', end_color="F2F2F2", fill_type='solid')
for row_number in range(1, active.max_row):
    if row_number % 2 == 0:
        for cell in active[row_number]:
            cell.fill = alternating_row_colour

#Resizing Columns to fit data
column_widths = []
for row in data:
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]
for i, column_width in enumerate(column_widths):
    active.column_dimensions[get_column_letter(i+1)].width = column_width

# Save workbook to write
wb.save(export_filename)


# # Generate PDF
# pdf = SimpleDocTemplate(
#     "outputPDF.pdf",
#     pagesize=letter
# )
#
# table = Table(data, hAlign='LEFT')
#
# # Table Styling
# style = TableStyle([
#     ('BACKGROUND', (0,0), (-1,0), Color( 176/256, 176/256, 176/256, alpha=1)),
#     ('TEXTCOLOR', (0,0), (-1,0), Color(255,255,255)),
#
#     ('ALIGN', (0,0), (-1,-1),'CENTER'),
#
#     ('FONTSIZE', (0,0), (-1,0), 13),
#     ('FONTSIZE', (0,1), (-1,-1), 11),
#
#     ('BOTTOMPADDING', (0,0,), (-1,0), 10),
#     ('TOPPADDING', (0,0,), (-1,0), 5),
#     ('LEFTPADDING', (0,0,), (-1,0), 10),
#     ('RIGHTPADDING', (0,0,), (-1,0), 10)
# ])
# table.setStyle(style)
#
# # Alternating Row Styling
# rowNum = len(data)
# for i in range(1, rowNum):
#     if i % 2 == 0:
#         bc = Color( 227/256, 227/256, 227/256, alpha=1)
#     else:
#         bc = Color( 245/256, 245/256, 245/256, alpha=1)
#
#     ts = TableStyle(
#         [('BACKGROUND', (0,i), (-1,i), bc)]
#     )
#     table.setStyle(ts)
#
# # Add datetime to PDF
# datetime = ('%4d/%02d/%02d %02d:%02d:%02d' % (
#  datetime.now().year,
#  datetime.now().month,
#  datetime.now().day,
#  datetime.now().hour,
#  datetime.now().minute,
#  datetime.now().second
# ))
#
# elems = []
# elems.append(Paragraph(datetime))
# elems.append(table)
#
# # Build PDF
# pdf.build(elems)
